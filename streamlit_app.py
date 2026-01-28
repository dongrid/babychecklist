import streamlit as st
from datetime import datetime, date, timedelta
import plotly.graph_objects as go
import openpyxl
import math

st.set_page_config(
    page_title="新生児管理チェックリスト",
    page_icon="👶",
    layout="wide"
)

st.title("新生児管理チェックリスト")
st.markdown("---")

st.markdown(
    """
<style>
div[role=\"radiogroup\"] {
  display: flex;
  gap: 0.5rem;
}
div[role=\"radiogroup\"] > label {
  border: 1px solid rgba(255,255,255,0.18);
  border-radius: 12px;
  padding: 0.35rem 0.75rem;
  margin: 0 !important;
  background: rgba(255,255,255,0.03);
}
div[role=\"radiogroup\"] > label:has(input:checked) {
  border-color: rgba(255,255,255,0.40);
  background: rgba(255,255,255,0.12);
}
div[role=\"radiogroup\"] > label > div {
  gap: 0.25rem;
}
div[role=\"radiogroup\"] svg {
  display: none;
}
</style>
""",
    unsafe_allow_html=True,
)

ICON_AABR = "👂"
ICON_MRI = "🧠"
ICON_EYE = "👁️"
ICON_THYROID = "🦋"
ICON_HYPOGLYCEMIA = "🩸"
ICON_JAUNDICE = "💡"


def lms_to_value(L, M, S, z):
    if L is None or M is None or S is None:
        return None
    if L == 0:
        return M * (2.718281828459045 ** (S * z))
    return M * ((1 + L * S * z) ** (1 / L))


def value_to_lms_z(L, M, S, value):
    if L is None or M is None or S is None or value is None:
        return None
    if value <= 0 or M <= 0 or S == 0:
        return None
    if L == 0:
        return math.log(value / M) / S
    return (((value / M) ** L) - 1) / (L * S)


def z_to_percentile(z):
    if z is None:
        return None
    return 50.0 * (1.0 + math.erf(z / math.sqrt(2.0)))


@st.cache_data(show_spinner=False)
def load_taikaku_birth_lms(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    sheet_name = wb.sheetnames[7]
    sh = wb[sheet_name]

    rows = {}
    last_week = None
    for r in range(7, sh.max_row + 1):
        week = sh.cell(r, 2).value
        day = sh.cell(r, 3).value
        if week is None:
            week = last_week
        if week is None or day is None:
            continue

        last_week = int(week)

        key = (int(week), int(day))
        rows[key] = {
            "maleFB_w": (sh.cell(r, 4).value, sh.cell(r, 5).value, sh.cell(r, 6).value),
            "maleSB_w": (sh.cell(r, 10).value, sh.cell(r, 11).value, sh.cell(r, 12).value),
            "femaleFB_w": (sh.cell(r, 16).value, sh.cell(r, 17).value, sh.cell(r, 18).value),
            "femaleSB_w": (sh.cell(r, 22).value, sh.cell(r, 23).value, sh.cell(r, 24).value),
            "birthH": (sh.cell(r, 34).value, sh.cell(r, 35).value, sh.cell(r, 36).value),
            # 頭囲LMS（Cols 28-30、性別・出生順位に関わらず共通）
            "maleFB_hc": (sh.cell(r, 28).value, sh.cell(r, 29).value, sh.cell(r, 30).value),
            "maleSB_hc": (sh.cell(r, 28).value, sh.cell(r, 29).value, sh.cell(r, 30).value),
            "femaleFB_hc": (sh.cell(r, 28).value, sh.cell(r, 29).value, sh.cell(r, 30).value),
            "femaleSB_hc": (sh.cell(r, 28).value, sh.cell(r, 29).value, sh.cell(r, 30).value),
        }

    return rows


def get_birth_size_thresholds(taikaku_rows, gender, is_first_child_bool, gestational_weeks, gestational_days):
    key = (int(gestational_weeks), int(gestational_days))
    row = taikaku_rows.get(key)
    if row is None:
        key = (int(gestational_weeks), 0)
        row = taikaku_rows.get(key)
    if row is None:
        w = int(gestational_weeks)
        for dw in range(1, 8):
            row = taikaku_rows.get((w - dw, 0))
            if row is not None:
                break
            row = taikaku_rows.get((w + dw, 0))
            if row is not None:
                break
    if row is None:
        return None

    if gender == "男児":
        weight_key = "maleFB_w" if is_first_child_bool else "maleSB_w"
        hc_key = "maleFB_hc" if is_first_child_bool else "maleSB_hc"
    else:
        weight_key = "femaleFB_w" if is_first_child_bool else "femaleSB_w"
        hc_key = "femaleFB_hc" if is_first_child_bool else "femaleSB_hc"

    wL, wM, wS = row[weight_key]
    hL, hM, hS = row["birthH"]
    hcL, hcM, hcS = row.get(hc_key, (None, None, None))

    z10 = -1.281551565545
    z90 = 1.281551565545
    z_minus2 = -2.0

    return {
        "weight_p10_g": lms_to_value(wL, wM, wS, z10),
        "weight_p90_g": lms_to_value(wL, wM, wS, z90),
        "weight_minus2sd_g": lms_to_value(wL, wM, wS, z_minus2),
        "height_p10_cm": lms_to_value(hL, hM, hS, z10),
        "height_p90_cm": lms_to_value(hL, hM, hS, z90),
        "height_minus2sd_cm": lms_to_value(hL, hM, hS, z_minus2),
        "hc_p10_cm": lms_to_value(hcL, hcM, hcS, z10),
        "hc_p90_cm": lms_to_value(hcL, hcM, hcS, z90),
        "hc_minus2sd_cm": lms_to_value(hcL, hcM, hcS, z_minus2),
        "weight_lms": (wL, wM, wS),
        "height_lms": (hL, hM, hS),
        "hc_lms": (hcL, hcM, hcS),
    }


def build_birth_size_plane_fig(birth_weight_g, birth_length_cm, thresholds):
    if thresholds is None:
        return None

    wx1 = thresholds.get("weight_minus2sd_g")
    wx2 = thresholds.get("weight_p10_g")
    wx3 = thresholds.get("weight_p90_g")
    hy1 = thresholds.get("height_minus2sd_cm")
    hy2 = thresholds.get("height_p10_cm")
    hy3 = thresholds.get("height_p90_cm")

    if None in (wx1, wx2, wx3, hy1, hy2, hy3):
        return None

    # 16領域：
    # - 体重：-2SD/10/90%ileで4区分
    # - 身長：-2SD/10/90%ileで4区分
    x_cuts = [wx1, wx2, wx3]
    y_cuts = [hy1, hy2, hy3]

    def bin_index(value, cuts):
        # cuts: [c0, c1, ...]
        for i, c in enumerate(cuts):
            if value < c:
                return i
        return len(cuts)

    xi = bin_index(birth_weight_g, x_cuts) if birth_weight_g is not None else 1
    yi = bin_index(birth_length_cm, y_cuts) if birth_length_cm is not None else 1

    # 表示範囲は「閾値と本児の周辺」にトリミングして見やすく
    if birth_weight_g is None:
        x_min = x_cuts[0] * 0.90
        x_max = x_cuts[-1] * 1.12
    else:
        x_min = min(birth_weight_g, x_cuts[0]) * 0.90
        x_max = max(birth_weight_g, x_cuts[-1]) * 1.12

    if birth_length_cm is None:
        y_min = y_cuts[0] * 0.98
        y_max = y_cuts[-1] * 1.04
    else:
        y_min = min(birth_length_cm, y_cuts[0]) * 0.98
        y_max = max(birth_length_cm, y_cuts[-1]) * 1.04

    x_edges = [x_min, x_cuts[0], x_cuts[1], x_cuts[2], x_max]
    y_edges = [y_min, y_cuts[0], y_cuts[1], y_cuts[2], y_max]

    x0, x1 = x_edges[xi], x_edges[xi + 1]
    y0, y1 = y_edges[yi], y_edges[yi + 1]

    fig = go.Figure()

    # 5分類の塗り分け（重なりが起きないように定義）
    # - HFD/LGA: 体重 > 90%ile
    # - AGA: 体重 10〜90%ile
    # - LFD: 体重 < 10%ile かつ 身長 >= 10%ile
    # - SGA: 体重 < 10%ile かつ 身長 < 10%ile
    # - SGA(GH): SGAのうち (体重 < -2SD または 身長 < -2SD)

    # HFD/LGA（右側帯）
    fig.add_shape(
        type="rect",
        x0=wx3,
        x1=x_edges[-1],
        y0=y_edges[0],
        y1=y_edges[-1],
        fillcolor="rgba(255, 120, 180, 0.14)",
        line=dict(color="rgba(255, 120, 180, 0.0)", width=0),
        layer="below",
    )

    # AGA（中央帯）はニュートラル：背景塗りはしない

    # 左側（体重<10%ile）を身長で分割：上=LFD、下=SGA（このあとSGA(GH)で一部上書き）
    # LFD（体重<10%ileの帯全体を指す概念。SGAを内包する）
    fig.add_shape(
        type="rect",
        x0=x_edges[0],
        x1=wx2,
        y0=y_edges[0],
        y1=y_edges[-1],
        fillcolor="rgba(255, 220, 120, 0.08)",
        line=dict(color="rgba(255, 220, 120, 0.0)", width=0),
        layer="below",
    )

    # SGA（ベース）: LFD帯のうち身長<10%ileの部分
    fig.add_shape(
        type="rect",
        x0=x_edges[0],
        x1=wx2,
        y0=y_edges[0],
        y1=hy2,
        fillcolor="rgba(110, 230, 160, 0.12)",
        line=dict(color="rgba(110, 230, 160, 0.0)", width=0),
        layer="below",
    )

    # SGA(GH)（SGAのうちL字領域）：SGAより強い同系色
    # - 左列：体重<-2SD & 身長<10%ile
    fig.add_shape(
        type="rect",
        x0=x_edges[0],
        x1=wx1,
        y0=y_edges[0],
        y1=hy2,
        fillcolor="rgba(110, 230, 160, 0.28)",
        line=dict(color="rgba(110, 230, 160, 0.0)", width=0),
        layer="below",
    )
    # - 下段：身長<-2SD & 体重が -2SD〜10%ile
    fig.add_shape(
        type="rect",
        x0=wx1,
        x1=wx2,
        y0=y_edges[0],
        y1=hy1,
        fillcolor="rgba(110, 230, 160, 0.28)",
        line=dict(color="rgba(110, 230, 160, 0.0)", width=0),
        layer="below",
    )

    # エリア内ラベル
    def mid(a, b):
        return a + (b - a) * 0.5

    label_font = dict(size=12, color="rgba(255,255,255,0.80)")
    box_bg = "rgba(0,0,0,0.22)"

    fig.add_annotation(
        x=mid(wx3, x_edges[-1]),
        y=mid(hy2, y_edges[-1]),
        text="HFD / LGA",
        showarrow=False,
        font=label_font,
        bgcolor=box_bg,
    )
    fig.add_annotation(
        x=mid(wx2, wx3),
        y=mid(hy2, y_edges[-1]),
        text="AGA",
        showarrow=False,
        font=label_font,
        bgcolor=box_bg,
    )
    fig.add_annotation(
        x=mid(x_edges[0], wx2),
        y=mid(hy2, y_edges[-1]),
        text="LFD",
        showarrow=False,
        font=label_font,
        bgcolor=box_bg,
    )
    fig.add_annotation(
        x=mid(wx1, wx2),
        y=mid(hy1, hy2),
        text="SGA",
        showarrow=False,
        font=label_font,
        bgcolor=box_bg,
    )
    fig.add_annotation(
        x=mid(x_edges[0], wx1),
        y=mid(y_edges[0], hy1),
        text="SGA (GH)",
        showarrow=False,
        font=label_font,
        bgcolor=box_bg,
    )

    line_color = "rgba(255,255,255,0.35)"
    label_color = "rgba(255,255,255,0.70)"
    dash_dot = "dot"

    # 線の近くにラベルを配置（凡例は使わない）
    y_label = y_edges[0] + (y_edges[-1] - y_edges[0]) * 0.03
    x_label = x_edges[0] + (x_edges[-1] - x_edges[0]) * 0.02

    for x, label, dash in [
        (wx1, f"体重 -2SD ({wx1:.0f}g)", dash_dot),
        (wx2, f"体重 10%ile ({wx2:.0f}g)", None),
        (wx3, f"体重 90%ile ({wx3:.0f}g)", None),
    ]:
        fig.add_shape(
            type="line",
            x0=x,
            x1=x,
            y0=y_edges[0],
            y1=y_edges[-1],
            line=dict(color=line_color, width=1, dash=dash),
            layer="below",
        )
        fig.add_annotation(
            x=x,
            y=y_label,
            text=label,
            showarrow=False,
            textangle=90,
            xanchor="left",
            yanchor="bottom",
            font=dict(size=11, color=label_color),
            bgcolor="rgba(0,0,0,0.25)",
        )

    for y, label, dash in [
        (hy1, f"身長 -2SD ({hy1:.1f}cm)", dash_dot),
        (hy2, f"身長 10%ile ({hy2:.1f}cm)", None),
        (hy3, f"身長 90%ile ({hy3:.1f}cm)", None),
    ]:
        fig.add_shape(
            type="line",
            x0=x_edges[0],
            x1=x_edges[-1],
            y0=y,
            y1=y,
            line=dict(color=line_color, width=1, dash=dash),
            layer="below",
        )
        fig.add_annotation(
            x=x_label,
            y=y,
            text=label,
            showarrow=False,
            xanchor="left",
            yanchor="bottom",
            font=dict(size=11, color=label_color),
            bgcolor="rgba(0,0,0,0.25)",
        )

    if birth_length_cm is None and birth_weight_g is not None:
        wL, wM, wS = thresholds.get("weight_lms", (None, None, None))
        w_z = value_to_lms_z(wL, wM, wS, birth_weight_g)
        w_p = z_to_percentile(w_z)
        w_z_text = "-" if w_z is None else f"{w_z:+.2f}SD"
        w_p_text = "-" if w_p is None else f"{w_p:.1f}%ile"

        fig.add_shape(
            type="line",
            x0=birth_weight_g,
            x1=birth_weight_g,
            y0=y_edges[0],
            y1=y_edges[-1],
            line=dict(color="rgba(255, 77, 77, 0.9)", width=3.5),
            layer="above",
        )
        fig.add_annotation(
            x=birth_weight_g,
            y=y_edges[-1],
            text=f"{birth_weight_g:.0f}g",
            showarrow=False,
            yanchor="bottom",
            xanchor="center",
            font=dict(size=11, color="rgba(255,255,255,0.8)"),
            bgcolor="rgba(0,0,0,0.25)",
        )

    if birth_weight_g is None and birth_length_cm is not None:
        fig.add_shape(
            type="line",
            x0=x_edges[0],
            x1=x_edges[-1],
            y0=birth_length_cm,
            y1=birth_length_cm,
            line=dict(color="rgba(255, 77, 77, 0.9)", width=3.5),
            layer="above",
        )
        fig.add_annotation(
            x=x_edges[-1],
            y=birth_length_cm,
            text=f"{birth_length_cm:.1f}cm",
            showarrow=False,
            yanchor="bottom",
            xanchor="right",
            font=dict(size=11, color="rgba(255,255,255,0.8)"),
            bgcolor="rgba(0,0,0,0.25)",
        )

    if birth_length_cm is not None and birth_weight_g is not None:
        wL, wM, wS = thresholds.get("weight_lms", (None, None, None))
        hL, hM, hS = thresholds.get("height_lms", (None, None, None))
        w_z = value_to_lms_z(wL, wM, wS, birth_weight_g)
        w_p = z_to_percentile(w_z)
        h_z = value_to_lms_z(hL, hM, hS, birth_length_cm)
        h_p = z_to_percentile(h_z)
        w_z_text = "-" if w_z is None else f"{w_z:+.2f}SD"
        w_p_text = "-" if w_p is None else f"{w_p:.1f}%ile"
        h_z_text = "-" if h_z is None else f"{h_z:+.2f}SD"
        h_p_text = "-" if h_p is None else f"{h_p:.1f}%ile"

        fig.add_trace(
            go.Scatter(
                x=[birth_weight_g],
                y=[birth_length_cm],
                mode="markers+text",
                text=[f"{birth_weight_g:.0f}g {birth_length_cm:.1f}cm"],
                textposition="top center",
                marker=dict(size=12, color="#ff4d4d", line=dict(color="rgba(0,0,0,0.5)", width=1)),
                hovertemplate="体重=%{x:.0f}g<br>身長=%{y:.1f}cm<extra></extra>",
                showlegend=False,
            )
        )


    fig.update_layout(
        margin=dict(l=10, r=10, t=40, b=10),
        xaxis_title="出生体重 (g)",
        yaxis_title="出生身長 (cm)",
        template="plotly_dark",
        height=380,
        showlegend=False,
    )
    fig.update_xaxes(range=[x_edges[0], x_edges[-1]])
    fig.update_yaxes(range=[y_edges[0], y_edges[-1]])
    return fig

# 村田基準の基準値（グローバル変数として定義）
MURATA_PHOTOTHERAPY_THRESHOLDS = {
    "≥ 2,500g": {
        0: 11.0, 1: 12.0, 2: 15.0, 3: 17.0,
        4: 18.0, 5: 19.0, 6: 19.5, 7: 20.0
    },
    "2,000 ~ 2,499g": {
        0: 9.5, 1: 10.0, 2: 12.0, 3: 14.0,
        4: 16.0, 5: 17.0, 6: 18.0, 7: 18.0
    },
    "1,500 ~ 1,999g": {
        0: 7.5, 1: 8.0, 2: 10.0, 3: 12.0,
        4: 14.0, 5: 15.0, 6: 16.0, 7: 16.0
    },
    "1,000 ~ 1,499g": {
        0: 6.5, 1: 7.0, 2: 7.0, 3: 8.0,
        4: 9.0, 5: 10.0, 6: 11.0, 7: 12.0
    },
    "≤ 999g": {
        0: 4.5, 1: 5.0, 2: 5.0, 3: 6.0,
        4: 7.0, 5: 8.0, 6: 9.0, 7: 10.0
    }
}

MORIOKA_TB_THRESHOLDS = {
    (22, 25): {
        24: (5, 6, 8),
        48: (5, 8, 10),
        72: (5, 8, 12),
        96: (6, 9, 13),
        120: (7, 10, 13),
        float("inf"): (8, 10, 13),
    },
    (26, 27): {
        24: (5, 6, 8),
        48: (5, 9, 10),
        72: (6, 9, 12),
        96: (8, 11, 14),
        120: (9, 12, 15),
        float("inf"): (10, 12, 15),
    },
    (28, 29): {
        24: (6, 7, 9),
        48: (7, 10, 12),
        72: (8, 12, 14),
        96: (10, 13, 16),
        120: (11, 14, 18),
        float("inf"): (12, 14, 18),
    },
    (30, 31): {
        24: (7, 8, 10),
        48: (8, 12, 14),
        72: (10, 14, 16),
        96: (12, 15, 18),
        120: (13, 16, 20),
        float("inf"): (14, 16, 20),
    },
    (32, 34): {
        24: (8, 9, 10),
        48: (10, 14, 16),
        72: (12, 16, 18),
        96: (14, 18, 20),
        120: (15, 19, 22),
        float("inf"): (16, 19, 22),
    },
    (35, float("inf")): {
        24: (10, 11, 12),
        48: (12, 16, 18),
        72: (14, 18, 20),
        96: (16, 20, 22),
        120: (17, 22, 25),
        float("inf"): (18, 22, 25),
    },
}

MORIOKA_UB_THRESHOLDS = {
    (22, 25): (0.4, 0.6, 0.8),
    (26, 27): (0.4, 0.6, 0.8),
    (28, 29): (0.5, 0.7, 0.9),
    (30, 31): (0.6, 0.8, 1.0),
    (32, 34): (0.7, 0.9, 1.2),
    (35, float("inf")): (0.8, 1.0, 1.5),
}


def get_morioka_thresholds(pca_weeks, hours_old):
    if pca_weeks is None or hours_old is None:
        return None

    pca_w = int(pca_weeks)
    group = None
    for (low, high) in MORIOKA_TB_THRESHOLDS.keys():
        if low <= pca_w <= high:
            group = (low, high)
            break
    if group is None:
        return None

    tb_bucket = None
    for upper_h in sorted(MORIOKA_TB_THRESHOLDS[group].keys(), key=lambda x: float(x)):
        if hours_old < upper_h:
            tb_bucket = upper_h
            break
    if tb_bucket is None:
        tb_bucket = float("inf")

    tb_low, tb_high, tb_exchange = MORIOKA_TB_THRESHOLDS[group][tb_bucket]
    ub_low, ub_high, ub_exchange = MORIOKA_UB_THRESHOLDS[group]

    if tb_bucket == float("inf"):
        time_label = "120時間以上"
    else:
        time_label = f"{int(tb_bucket)}時間未満"

    return {
        "pca_group": group,
        "time_bucket_hours": tb_bucket,
        "time_label": time_label,
        "tb": {"low": tb_low, "high": tb_high, "exchange": tb_exchange},
        "ub": {"low": ub_low, "high": ub_high, "exchange": ub_exchange},
    }


def build_morioka_html_table(
    current_pca_group=None,
    current_time_bucket_hours=None,
    highlight_pairs=None,
):
    time_buckets = [24, 48, 72, 96, 120, float("inf")]
    time_labels = {
        24: "<24h",
        48: "<48h",
        72: "<72h",
        96: "<96h",
        120: "<120h",
        float("inf"): "120h-",
    }

    table_style = "border-collapse:collapse;width:100%;font-size:14px"
    th_style = "text-align:center;padding:6px 8px;border:1px solid #333;background:#1b1b1b;color:#eaeaea;white-space:nowrap"
    th_left_style = "text-align:left;padding:6px 8px;border:1px solid #333;background:#1b1b1b;color:#eaeaea;white-space:nowrap"
    td_style = "padding:6px 8px;border:1px solid #333;background:#0f0f0f;color:#eaeaea;text-align:center"

    if highlight_pairs is None:
        highlight_pairs = set()

    rows = []
    for (pca_low, pca_high), tb_by_bucket in MORIOKA_TB_THRESHOLDS.items():
        if pca_high == float("inf"):
            pca_label = f"{pca_low}w-"
        else:
            pca_label = f"{pca_low}-{pca_high}w"

        ub_low, ub_high, ub_ex = MORIOKA_UB_THRESHOLDS[(pca_low, pca_high)]
        ub_cell = f"{ub_low}/{ub_high}/{ub_ex}"

        row_cells = []
        for b in time_buckets:
            low, high, ex = tb_by_bucket[b]
            cell = f"{low}/{high}/{ex}"
            is_cell_hit = (current_pca_group == (pca_low, pca_high) and current_time_bucket_hours == b)
            style = td_style
            is_soft_hit = ((pca_low, pca_high), b) in highlight_pairs
            if is_soft_hit:
                style = style + ";background:rgba(255, 238, 186, 0.18)"
            if is_cell_hit:
                style = style + ";background:#ffeeba;color:#111;font-weight:800"
            row_cells.append(f"<td style='{style}'>{cell}</td>")

        is_row_hit = (current_pca_group == (pca_low, pca_high))
        row_style = ""
        if is_row_hit:
            row_style = "background:#141414"

        ub_td_style = td_style
        if is_row_hit:
            ub_td_style = ub_td_style + ";background:#ffeeba;color:#111;font-weight:800"

        rows.append(
            "<tr style='" + row_style + "'>"
            + f"<th style='{th_left_style}'>{pca_label}</th>"
            + "".join(row_cells)
            + f"<td style='{ub_td_style}'>{ub_cell}</td>"
            + "</tr>"
        )

    header = (
        "<tr>"
        f"<th style='{th_left_style}'>修正週数</th>"
        + "".join(
            [
                f"<th style='{th_style}'>{time_labels[b]}</th>"
                for b in time_buckets
            ]
        )
        + f"<th style='{th_style}'>UB（µg/dL）</th>"
        + "</tr>"
    )

    return (
        "<div style='overflow-x:auto'>"
        f"<table style='{table_style}'>"
        + header
        + "".join(rows)
        + "</table>"
        "</div>"
    )


def get_morioka_pca_group_from_weeks(pca_weeks):
    if pca_weeks is None:
        return None
    w = int(pca_weeks)
    for (low, high) in MORIOKA_TB_THRESHOLDS.keys():
        if low <= w <= high:
            return (low, high)
    return None


def build_morioka_ub_html_table(current_pca_group=None):
    table_style = "border-collapse:collapse;width:100%;font-size:14px"
    th_style = "text-align:center;padding:6px 8px;border:1px solid #333;background:#1b1b1b;color:#eaeaea;white-space:nowrap"
    th_left_style = "text-align:left;padding:6px 8px;border:1px solid #333;background:#1b1b1b;color:#eaeaea;white-space:nowrap"
    td_style = "padding:6px 8px;border:1px solid #333;background:#0f0f0f;color:#eaeaea;text-align:center"

    header = (
        "<tr>"
        f"<th style='{th_left_style}'>修正週数</th>"
        f"<th style='{th_style}'>low</th>"
        f"<th style='{th_style}'>high</th>"
        f"<th style='{th_style}'>交換輸血</th>"
        "</tr>"
    )

    rows = []
    for (pca_low, pca_high), (ub_low, ub_high, ub_ex) in MORIOKA_UB_THRESHOLDS.items():
        if pca_high == float("inf"):
            pca_label = f"{pca_low}w-"
        else:
            pca_label = f"{pca_low}-{pca_high}w"

        is_row_hit = (current_pca_group == (pca_low, pca_high))
        hit_td_style = td_style + ";background:#ffeeba;color:#111;font-weight:800" if is_row_hit else td_style

        rows.append(
            "<tr>"
            + f"<th style='{th_left_style}'>{pca_label}</th>"
            + f"<td style='{hit_td_style}'>{ub_low}</td>"
            + f"<td style='{hit_td_style}'>{ub_high}</td>"
            + f"<td style='{hit_td_style}'>{ub_ex}</td>"
            + "</tr>"
        )

    return (
        "<div style='overflow-x:auto'>"
        f"<table style='{table_style}'>"
        + header
        + "".join(rows)
        + "</table>"
        "</div>"
    )

def get_phototherapy_threshold(weight, days_old, has_kernicterus_risk=False):
    """村田・井村の基準に基づいて光線療法基準値を取得"""
    
    # 出生体重カテゴリーの決定
    if weight >= 2500:
        category = "≥ 2,500g"
    elif weight >= 2000:
        category = "2,000 ~ 2,499g"
    elif weight >= 1500:
        category = "1,500 ~ 1,999g"
    elif weight >= 1000:
        category = "1,000 ~ 1,499g"
    else:  # weight < 1000
        category = "≤ 999g"
    
    original_category = category
    
    # 核黄疸危険因子がある場合は1段階低い基準を使用
    if has_kernicterus_risk:
        if category == "≥ 2,500g":
            category = "2,000 ~ 2,499g"
        elif category == "2,000 ~ 2,499g":
            category = "1,500 ~ 1,999g"
        elif category == "1,500 ~ 1,999g":
            category = "1,000 ~ 1,499g"
        elif category == "1,000 ~ 1,499g":
            category = "≤ 999g"
        # ≤ 999g の場合はこれ以上低い基準がないので、そのまま使用
    
    thresholds = MURATA_PHOTOTHERAPY_THRESHOLDS[category]
    
    # 日齢に応じた基準値を取得
    if days_old == 0:
        # 0日目は村田・井村の基準として定義がないため、数値を返さない
        day = 0
        threshold = None
        is_day0 = True
    else:
        day = min(days_old, 7)
        threshold = thresholds.get(day, thresholds[7])
        is_day0 = False
    
    # 0日目の基準は未定義のため、参考値も返さない
    day0_threshold = None
    
    # 核黄疸危険因子により基準を変更した場合の情報も返す
    adjusted = has_kernicterus_risk and original_category != category
    
    return category, threshold, adjusted, original_category, is_day0, day0_threshold

def get_management_guidance(weight, is_first_child, delivery_method, gestational_age, days_old,
                           maternal_diabetes=False, maternal_thyroid_abnormal=False,
                           apgar_score_5min=9, delivery_stress=False, birth_date=None, birth_time=None,
                           exchange_transfusion=False, intracranial_hemorrhage=False,
                           apnea_treatment=False, aminoglycoside_history=False,
                           high_oxygen=False, corrected_weeks=0,
                           gestational_weeks=0, gestational_days=0,
                           weight_lt_p10=False, weight_ge_p90=False):
    """新生児の体重や状況に基づいて管理方針を決定"""
    
    guidance = {
        'category': '',
        'recommendations': [],
        'warnings': [],
        'special_management': []
    }
    
    # 分類（出生体重 + 早産の程度）
    if weight >= 4000:
        weight_cat = '高出生体重児'
    elif weight >= 2500:
        weight_cat = '正常出生体重児'
    elif weight < 1000:
        weight_cat = '超極低出生体重児（ELBW）'
    elif weight < 1500:
        weight_cat = '極低出生体重児（VLBW）'
    else:
        weight_cat = '低出生体重児（LBW）'

    if gestational_age >= 42:
        prematurity_cat = '過期産'
    elif gestational_age >= 37:
        prematurity_cat = '正期産'
    elif gestational_age >= 34:
        prematurity_cat = '後期早産'
    else:
        prematurity_cat = '早産'

    guidance['category'] = f"{prematurity_cat} / {weight_cat}"
    
    # ケイツーシロップ12回投与法（すべての子どもに適応）
    k2_third_to_twelfth = None
    
    # 3回目以降：日齢11以降に迎える水曜日から毎週水曜日に12回目まで
    if birth_date and birth_time:
        birth_datetime = datetime.combine(birth_date, birth_time)
        # 日齢11以降に迎える最初の水曜日を計算
        first_wednesday_after_day11 = None
        for i in range(11, 18):  # 日齢11から17の間で最初の水曜日を探す
            check_date = birth_date + timedelta(days=i)
            if check_date.weekday() == 2:  # 水曜日
                first_wednesday_after_day11 = check_date
                break
        
        if first_wednesday_after_day11:
            last_wednesday = first_wednesday_after_day11 + timedelta(weeks=9)  # 12回目（3回目から10週後）
            k2_third_to_twelfth = f'{first_wednesday_after_day11.strftime("%Y/%m/%d")}から{last_wednesday.strftime("%Y/%m/%d")}まで毎週水曜日に内服'
    
    # すべての子どもに適応があるため、常に表示
    guidance['special_management'].append({
        'title': '💊 ケイツーシロップ12回投与法',
        'k2_third_to_twelfth': k2_third_to_twelfth,
        'items': [
            '・入院中の内服は処置オーダで指示する',
            '・退院処方として12回目までのケイツーを処方する'
        ],
        'needed': True
    })
    
    # マススクリーニング（すべての子どもに適応）
    mass_screening_items = []
    if birth_date:
        day4_date = birth_date + timedelta(days=4)
        mass_screening_items.append(f'・日齢4（{day4_date.strftime("%Y/%m/%d")}）：マススクリーニングを実施（希望あれば拡大マスも）')
    else:
        mass_screening_items.append('・日齢4：マススクリーニングを実施（希望あれば拡大マスも）')
    
    # 早産児は退院前にマススクリーニング再検
    if gestational_age < 37:
        mass_screening_items.append('・早産児のため、退院前にマススクリーニング再検を行う')
    
    # すべての子どもに適応があるため、常に表示
    guidance['special_management'].append({
        'title': '🧪 マススクリーニング',
        'items': mass_screening_items,
        'needed': True
    })
    
    # 血糖チェック
    hypoglycemia_risk = (
        gestational_age < 37 or
        weight < 2500 or
        maternal_diabetes or
        weight_ge_p90 or
        maternal_diabetes or
        delivery_stress or
        apgar_score_5min < 7
    )
    
    if hypoglycemia_risk:
        # 適応理由を取得
        risk_reasons = []
        if gestational_age < 37:
            risk_reasons.append("在胎37週未満")
        if weight < 2500:
            risk_reasons.append("出生体重2500g未満")
        if weight_ge_p90:
            risk_reasons.append("出生体重90%ile以上")
        if maternal_diabetes:
            risk_reasons.append("妊娠糖尿病")
        if delivery_stress:
            risk_reasons.append("分娩ストレス")
        if apgar_score_5min < 7:
            risk_reasons.append("Apgar5分値7未満")
        
        items = [
            '・出生後できるだけ早期にミルクを開始（糖水は避ける）',
            '・その後も3時間毎に哺乳を継続',
            '・出生3/6/12時間後に簡易血糖測定を実施'
        ]
        guidance['special_management'].append({
            'title': '🩸 血糖チェック',
            'items': items,
            'needed': True
        })
    else:
        # 適応がない理由を判定
        reason = []
        if gestational_age >= 37:
            reason.append("在胎37週以上")
        if weight >= 2500:
            reason.append("体重2500g以上")
        if not maternal_diabetes:
            reason.append("糖尿病母体なし")
        if not delivery_stress:
            reason.append("分娩ストレスなし")
        if apgar_score_5min >= 7:
            reason.append("Apgar5分値7以上")
        guidance['special_management'].append({
            'title': '🩸 血糖チェック',
            'items': [f'・適応なし（{"、".join(reason[:3]) if reason else "低血糖リスク因子なし"}）'],
            'needed': False
        })
    
    # 甲状腺機能検査の対象児
    thyroid_check_needed = maternal_thyroid_abnormal
    
    if thyroid_check_needed:
        thyroid_items = [
            '・適応理由：母体の甲状腺異常あり',
            '・日齢5でTSH/FT4を測定',
            '・必要に応じて小児科内分泌に相談'
        ]
        guidance['special_management'].append({
            'title': '🦋 甲状腺機能検査',
            'items': thyroid_items,
            'needed': True
        })
    else:
        guidance['special_management'].append({
            'title': '🦋 甲状腺機能検査',
            'items': ['・適応なし（母体の甲状腺関連情報なし）'],
            'needed': False
        })
    
    # 頭部MRIを実施する条件
    mri_needed = (
        gestational_age < 34 or
        weight < 1500 or
        exchange_transfusion or
        intracranial_hemorrhage
    )
    
    if mri_needed:
        # 適応理由を取得
        mri_reasons = []
        if gestational_age < 34:
            mri_reasons.append("在胎34週未満")
        if weight < 1500:
            mri_reasons.append("体重1500g未満")
        if exchange_transfusion:
            mri_reasons.append("交換輸血を実施")
        if intracranial_hemorrhage:
            mri_reasons.append("頭蓋内出血")
        
        mri_items = [
            f'・適応理由：{"、".join(mri_reasons)}',
            '・退院前に頭部MRIを実施',
            '・時期：全身状態が安定した頃'
        ]
        if weight < 1000:  # 極低出生体重児
            mri_items.append('・極低出生体重児は修正37-44週で検査時体重1500g以上')
        guidance['special_management'].append({
            'title': '🧠 頭部MRI',
            'items': mri_items,
            'needed': True
        })
    else:
        # 適応がない理由を判定
        reason = []
        if gestational_age >= 34:
            reason.append("在胎34週以上")
        if weight >= 1500:
            reason.append("体重1500g以上")
        if not exchange_transfusion:
            reason.append("交換輸血なし")
        if not intracranial_hemorrhage:
            reason.append("頭蓋内出血なし")
        guidance['special_management'].append({
            'title': '🧠 頭部MRI',
            'items': [f'・適応なし（{"、".join(reason[:2]) if reason else "適応条件を満たさない"}）'],
            'needed': False
        })
    
    # AABR
    aabr_insurance = (
        gestational_age < 35 or
        weight <= 1800 or
        exchange_transfusion or  # 重症黄疸（交換輸血を実施）
        apnea_treatment or
        aminoglycoside_history or
        intracranial_hemorrhage
    )
    
    if aabr_insurance:
        # 適応理由を取得
        aabr_reasons = []
        if gestational_age < 35:
            aabr_reasons.append("在胎35週未満")
        if weight <= 1800:
            aabr_reasons.append("体重1800g以下")
        if exchange_transfusion:
            aabr_reasons.append("交換輸血を実施")
        if apnea_treatment:
            aabr_reasons.append("無呼吸発作治療")
        if aminoglycoside_history:
            aabr_reasons.append("アミノグリコシド投与歴")
        if intracranial_hemorrhage:
            aabr_reasons.append("頭蓋内出血")
        
        guidance['special_management'].append({
            'title': '👂 AABR',
            'items': [
                f'・保険適応理由：{"、".join(aabr_reasons)}',
                '・時期：全身状態が安定した頃'
            ],
            'needed': True
        })
    else:
        # 適応がない理由を判定
        reason = []
        if gestational_age >= 35:
            reason.append("在胎35週以上")
        if weight > 1800:
            reason.append("体重1800g超")
        if not exchange_transfusion:
            reason.append("交換輸血なし")
        if not apnea_treatment:
            reason.append("無呼吸発作治療なし")
        if not aminoglycoside_history:
            reason.append("アミノグリコシド投与歴なし")
        if not intracranial_hemorrhage:
            reason.append("頭蓋内出血なし")
        guidance['special_management'].append({
            'title': '👂 AABR',
            'items': [f'・保険適応なし（{"、".join(reason[:2]) if reason else "適応条件を満たさない"}。ご家族の希望により自費で実施可能）'],
            'needed': False
        })
    
    # 眼底検査
    eye_exam_needed = (
        gestational_age < 34 or
        weight < 1800 or
        high_oxygen
    )
    
    if eye_exam_needed:
        # 適応理由を取得
        eye_reasons = []
        if gestational_age < 34:
            eye_reasons.append("在胎34週未満")
        if weight < 1800:
            eye_reasons.append("体重1800g未満")
        if high_oxygen:
            eye_reasons.append("高濃度酸素投与歴")
        
        eye_items = [
            f'・適応理由：{"、".join(eye_reasons)}',
            '・眼科に診察を依頼する',
            '・時期：生後2-3週毎',
            '・準備：サンドールP点眼液を事前に処方しておく',
            '・眼科宛の院内紹介状を作成し、眼科受診の指示をしておく',
            '・必要な場合には、眼科処置前後のミルク量を減らす指示を出しておく'
        ]
        guidance['special_management'].append({
            'title': '👁️ 眼底検査',
            'items': eye_items,
            'needed': True
        })
    else:
        # 適応がない理由を判定
        reason = []
        if gestational_age >= 34:
            reason.append("在胎34週以上")
        if weight >= 1800:
            reason.append("体重1800g以上")
        if not high_oxygen:
            reason.append("高濃度酸素投与歴なし")
        guidance['special_management'].append({
            'title': '👁️ 眼底検査',
            'items': [f'・適応なし（{"、".join(reason[:2]) if reason else "適応条件を満たさない"}）'],
            'needed': False
        })
    
    # 特別な管理項目をrecommendationsに変換
    for special in guidance['special_management']:
        if special.get('needed', True):
            guidance['recommendations'].append(f"**{special['title']}**")
            for item in special['items']:
                guidance['recommendations'].append(item)
        else:
            guidance['recommendations'].append(f"**{special['title']}**")
            guidance['recommendations'].append(f"<span style='color: gray;'>{special['items'][0]}</span>")
    
    return guidance

# 入力フィールド
st.header("✍️ 入力")

# 基本情報
st.subheader("ℹ️ 基本情報")
row1 = st.columns([1.8, 1.6, 1.2, 0.8])
with row1[0]:
    birth_date = st.date_input(
        "出生日",
        value=date.today(),
        max_value=date.today()
    )
with row1[1]:
    birth_time = st.time_input("出生時刻", value=datetime.now().time())
with row1[2]:
    gestational_weeks = st.number_input(
        "在胎週数（週）",
        min_value=20,
        max_value=42,
        value=39,
        step=1
    )
with row1[3]:
    gestational_days = st.number_input(
        "（日）",
        min_value=0,
        max_value=6,
        value=0,
        step=1
    )

row2 = st.columns([1.5, 1.5, 1.5, 1.5])
with row2[0]:
    col_wt, col_wt_unknown = st.columns([3, 1], vertical_alignment="bottom")
    with col_wt:
        birth_weight = st.number_input(
            "出生体重 (g)",
            min_value=500,
            max_value=6000,
            value=3000,
            step=1,
        )
    with col_wt_unknown:
        birth_weight_unknown = st.checkbox("未測定", key="birth_weight_unknown")

with row2[1]:
    col_len, col_len_unknown = st.columns([3, 1], vertical_alignment="bottom")
    with col_len:
        birth_length = st.number_input(
            "出生身長 (cm)",
            min_value=20.0,
            max_value=70.0,
            value=50.0,
            step=0.1,
            format="%.1f",
        )
    with col_len_unknown:
        birth_length_unknown = st.checkbox("未測定", key="birth_length_unknown")

with row2[2]:
    col_hc, col_hc_unknown = st.columns([3, 1], vertical_alignment="bottom")
    with col_hc:
        birth_head_circumference = st.number_input(
            "出生頭囲 (cm)",
            min_value=15.0,
            max_value=45.0,
            value=33.5,
            step=0.1,
            format="%.1f",
        )
    with col_hc_unknown:
        birth_head_circumference_unknown = st.checkbox("未測定", key="birth_head_circumference_unknown")

with row2[3]:
    gender = st.radio(
        "性別",
        ["男児", "女児"],
        horizontal=True,
    )

row2_5 = st.columns([1.5, 1.5, 1.5, 1.5])
with row2_5[0]:
    is_first_child = st.radio(
        "出生順位",
        ["初産", "経産"],
        horizontal=True,
    )

# 値の変換処理
birth_weight = None if birth_weight_unknown else float(birth_weight)
birth_length = None if birth_length_unknown else float(birth_length)
birth_head_circumference = None if birth_head_circumference_unknown else float(birth_head_circumference)

if birth_weight_unknown:
    st.markdown(
        """
<style>
div[data-testid="stNumberInput"] input[aria-label="出生体重 (g)"] {
  opacity: 0.45;
}
</style>
""",
        unsafe_allow_html=True,
    )

if birth_length_unknown:
    st.markdown(
        """
<style>
div[data-testid="stNumberInput"] input[aria-label="出生身長 (cm)"] {
  opacity: 0.45;
}
</style>
""",
        unsafe_allow_html=True,
    )

if birth_head_circumference_unknown:
    st.markdown(
        """
<style>
div[data-testid="stNumberInput"] input[aria-label="出生頭囲 (cm)"] {
  opacity: 0.45;
}
</style>
""",
        unsafe_allow_html=True,
    )

# 分娩情報 + Apgar
row3 = st.columns([1.8, 1, 1])
with row3[0]:
    delivery_method = st.selectbox(
        "分娩形式",
        ["経腟分娩", "計画帝王切開", "緊急帝王切開", "吸引・鉗子分娩", "その他"]
    )
with row3[1]:
    apgar_score_1min = st.number_input(
        "Apgar（1分）",
        min_value=0,
        max_value=10,
        value=9,
        step=1
    )
with row3[2]:
    apgar_score_5min = st.number_input(
        "Apgar（5分）",
        min_value=0,
        max_value=10,
        value=9,
        step=1
    )
gestational_age = gestational_weeks + gestational_days / 7.0


# 追加情報
st.subheader("🤰 母体情報")
col1, col2 = st.columns(2)
with col1:
    maternal_diabetes = st.checkbox(f"妊娠糖尿病 {ICON_HYPOGLYCEMIA}")
with col2:
    maternal_thyroid_abnormal = st.checkbox(f"甲状腺異常 {ICON_THYROID}")
st.caption("甲状腺異常：内服加療中 / 抗甲状腺抗体陽性 / 既往あり（妊娠経過の情報不明）などを含む")

st.subheader("👶 新生児情報")
col1, col2 = st.columns(2)
with col1:
    exchange_transfusion = st.checkbox(f"重症黄疸（交換輸血を実施） {ICON_JAUNDICE}{ICON_AABR}{ICON_MRI}")
    intracranial_hemorrhage = st.checkbox(f"頭蓋内出血 {ICON_AABR}{ICON_MRI}")
    apnea_treatment = st.checkbox(f"無呼吸発作治療 {ICON_AABR}")
    high_oxygen = st.checkbox(f"高濃度酸素投与歴 {ICON_EYE}")
    respiratory_distress = st.checkbox(f"呼吸窮迫（PaO2≦40が2時間以上持続） {ICON_JAUNDICE}")
    acidosis = st.checkbox(f"アシドーシス（pH≦7.15） {ICON_JAUNDICE}")
with col2:
    aminoglycoside_history = st.checkbox(f"アミノグリコシド投与歴（ゲンタシン/アミカシンなど） {ICON_AABR}")
    hypothermia = st.checkbox(f"低体温（直腸温<35℃が2時間以上持続） {ICON_JAUNDICE}")
    hypoproteinemia = st.checkbox(f"低蛋白血症（血清蛋白≦4.0またはAlb≦2.5） {ICON_JAUNDICE}")
    hypoglycemia = st.checkbox(f"低血糖 {ICON_JAUNDICE}")
    hemolysis = st.checkbox(f"溶血 {ICON_JAUNDICE}")
    cns_abnormality = st.checkbox(f"敗血症を含む中枢神経系の異常徴候 {ICON_JAUNDICE}")

# 日齢の計算
today = date.today()
days_old = (today - birth_date).days

now_dt = datetime.now()
birth_dt = datetime.combine(birth_date, birth_time)
hours_old = (now_dt - birth_dt).total_seconds() / 3600
if hours_old < 0:
    hours_old = 0.0

# 修正週数・日数の計算
total_gestational_days = gestational_weeks * 7 + gestational_days
corrected_total_days = total_gestational_days + days_old
corrected_weeks = corrected_total_days // 7
corrected_days = corrected_total_days % 7

morioka = get_morioka_thresholds(corrected_weeks, hours_old)

# 核黄疸危険因子の自動判定（Apgarスコア5分値≦3、またはその他の危険因子がある場合）
has_kernicterus_risk = (
    apgar_score_5min <= 3 or
    respiratory_distress or
    acidosis or
    hypothermia or
    hypoproteinemia or
    hypoglycemia or
    hemolysis or
    cns_abnormality
)

# 光線療法基準の計算
if birth_weight is None:
    phototherapy_category = None
    phototherapy_threshold = None
    adjusted = False
    original_category = None
    is_day0 = False
    day0_threshold = None
else:
    phototherapy_category, phototherapy_threshold, adjusted, original_category, is_day0, day0_threshold = get_phototherapy_threshold(
        birth_weight,
        days_old,
        has_kernicterus_risk
    )

st.markdown("---")
st.header("🏷️ 判定結果")

# 日齢と修正週数・日数の表示
col1, col2 = st.columns(2)
with col1:
    st.metric("日齢（今日）", f"{days_old} 日")
with col2:
    st.metric("修正週数・日数（今日）", f"{corrected_weeks}週{corrected_days}日")

# 分娩ストレスの判定
delivery_stress = (
    delivery_method in ["吸引・鉗子分娩", "緊急帝王切開"] or
    apgar_score_5min < 7
)

# 管理方針の取得
is_first_child_bool = is_first_child == "初産"

taikaku_rows = load_taikaku_birth_lms("taikakubirthlongcross_v1.1.xlsx")
birth_thresholds = get_birth_size_thresholds(
    taikaku_rows,
    gender,
    is_first_child_bool,
    gestational_weeks,
    gestational_days,
)

weight_lt_p10 = False
weight_ge_p90 = False
weight_lt_minus2sd = False
height_lt_p10 = False
height_lt_minus2sd = False
if birth_thresholds is not None:
    if birth_weight is not None and birth_thresholds.get("weight_p10_g") is not None:
        weight_lt_p10 = birth_weight < birth_thresholds["weight_p10_g"]
    if birth_weight is not None and birth_thresholds.get("weight_p90_g") is not None:
        weight_ge_p90 = birth_weight > birth_thresholds["weight_p90_g"]
    if birth_weight is not None and birth_thresholds.get("weight_minus2sd_g") is not None:
        weight_lt_minus2sd = birth_weight < birth_thresholds["weight_minus2sd_g"]

    if birth_length is not None:
        if birth_thresholds.get("height_p10_cm") is not None:
            height_lt_p10 = birth_length < birth_thresholds["height_p10_cm"]
        if birth_thresholds.get("height_minus2sd_cm") is not None:
            height_lt_minus2sd = birth_length < birth_thresholds["height_minus2sd_cm"]

eu_sga = False
lfd = False
sga = False
sga_gh = False
aga = False
hfd_lga = False

if birth_weight is None:
    eu_sga = False
    lfd = False
    sga = False
    sga_gh = False
    aga = False
    hfd_lga = False
elif birth_length is None:
    eu_sga = weight_lt_minus2sd
    # 身長未測定時：体重のみで判定
    lfd = weight_lt_p10
    aga = (not weight_lt_p10) and (not weight_ge_p90)
    hfd_lga = weight_ge_p90
else:
    eu_sga = (weight_lt_minus2sd or height_lt_minus2sd)
    lfd = (weight_lt_p10 and (not height_lt_p10))
    sga = (weight_lt_p10 and height_lt_p10)
    sga_gh = (sga and (weight_lt_minus2sd or height_lt_minus2sd))
    aga = (not weight_lt_p10) and (not weight_ge_p90)
    hfd_lga = weight_ge_p90

birth_plane_fig = build_birth_size_plane_fig(birth_weight, birth_length, birth_thresholds)
if birth_weight is None:
    if gestational_age >= 42:
        prematurity_cat = "過期産"
    elif gestational_age >= 37:
        prematurity_cat = "正期産"
    elif gestational_age >= 34:
        prematurity_cat = "後期早産"
    else:
        prematurity_cat = "早産"

    mass_screening_items = []
    if birth_date:
        day4_date = birth_date + timedelta(days=4)
        mass_screening_items.append(
            f"・日齢4（{day4_date.strftime('%Y/%m/%d')}）：マススクリーニングを実施（希望あれば拡大マスも）"
        )
    else:
        mass_screening_items.append("・日齢4：マススクリーニングを実施（希望あれば拡大マスも）")
    if gestational_age < 37:
        mass_screening_items.append("・早産児のため、退院前にマススクリーニング再検を行う")

    guidance = {
        "category": prematurity_cat,
        "recommendations": [],
        "warnings": [],
        "special_management": [
            {
                "title": "💊 ケイツーシロップ12回投与法",
                "k2_third_to_twelfth": None,
                "items": [
                    "・入院中の内服は処置オーダで指示する",
                    "・退院処方として12回目までのケイツーを処方する",
                ],
                "needed": True,
            },
            {
                "title": "🧪 マススクリーニング",
                "items": mass_screening_items,
                "needed": True,
            },
            {
                "title": "⚠️ 体重未測定のため一部判定不可",
                "items": [
                    "・出生体重区分、低血糖リスク、SGA/LGA判定などは体重測定後に評価",
                ],
                "needed": False,
            },
        ],
    }
else:
    guidance = get_management_guidance(
        birth_weight,
        is_first_child_bool,
        delivery_method,
        gestational_age,
        days_old,
        maternal_diabetes,
        maternal_thyroid_abnormal,
        apgar_score_5min,
        delivery_stress,
        birth_date,
        birth_time,
        exchange_transfusion,
        intracranial_hemorrhage,
        apnea_treatment,
        aminoglycoside_history,
        high_oxygen,
        corrected_weeks,
        gestational_weeks,
        gestational_days,
        weight_lt_p10,
        weight_ge_p90
    )

# 分類の表示
if birth_thresholds is None:
    birth_size_label = None
elif birth_weight is None:
    birth_size_label = None
else:
    if hfd_lga:
        birth_size_label = "heavy-for-dates (HFD) / large for gestational age (LGA)"
    elif aga:
        birth_size_label = "appropriate for gestational age (AGA)"
    elif sga_gh:
        birth_size_label = "small for gestational age (SGA)（GH適応）"
    elif sga:
        birth_size_label = "small for gestational age (SGA)（GH適応なし）"
    elif lfd:
        birth_size_label = "light-for-dates (LFD)"
    else:
        birth_size_label = "判定不可"
if birth_size_label:
    st.subheader(f"分類: {guidance['category']} / {birth_size_label}")
else:
    st.subheader(f"分類: {guidance['category']}")

if birth_thresholds is not None:
    st.caption(f"{gender} / {is_first_child} / 在胎{gestational_weeks}週{gestational_days}日")

    wL, wM, wS = birth_thresholds.get("weight_lms", (None, None, None))
    w_z = value_to_lms_z(wL, wM, wS, birth_weight)
    w_p = z_to_percentile(w_z)

    w_z_text = "-" if w_z is None else f"{w_z:+.2f}SD"
    w_p_text = "-" if w_p is None else f"{w_p:.1f}%ile"
    if birth_weight is None:
        st.markdown(
            "体重: **未測定**"
            f"（-2SD {birth_thresholds['weight_minus2sd_g']:.0f}g / 10%ile {birth_thresholds['weight_p10_g']:.0f}g / 90%ile {birth_thresholds['weight_p90_g']:.0f}g）"
        )
    else:
        st.markdown(
            f"体重: **{birth_weight:.0f}g / {w_z_text} / {w_p_text}**"
            f"（-2SD {birth_thresholds['weight_minus2sd_g']:.0f}g / 10%ile {birth_thresholds['weight_p10_g']:.0f}g / 90%ile {birth_thresholds['weight_p90_g']:.0f}g）"
        )

    if birth_length is None:
        st.markdown(
            "身長: **未測定**"
            f"（-2SD {birth_thresholds['height_minus2sd_cm']:.1f}cm / 10%ile {birth_thresholds['height_p10_cm']:.1f}cm / 90%ile {birth_thresholds['height_p90_cm']:.1f}cm）"
        )
    else:
        hL, hM, hS = birth_thresholds.get("height_lms", (None, None, None))
        h_z = value_to_lms_z(hL, hM, hS, birth_length)
        h_p = z_to_percentile(h_z)
        h_z_text = "-" if h_z is None else f"{h_z:+.2f}SD"
        h_p_text = "-" if h_p is None else f"{h_p:.1f}%ile"
        st.markdown(
            f"身長: **{birth_length:.1f}cm / {h_z_text} / {h_p_text}**"
            f"（-2SD {birth_thresholds['height_minus2sd_cm']:.1f}cm / 10%ile {birth_thresholds['height_p10_cm']:.1f}cm / 90%ile {birth_thresholds['height_p90_cm']:.1f}cm）"
        )
    
    if birth_head_circumference is None:
        if birth_thresholds.get("hc_p10_cm") is not None:
            st.markdown(
                "頭囲: **未測定**"
                f"（-2SD {birth_thresholds['hc_minus2sd_cm']:.1f}cm / 10%ile {birth_thresholds['hc_p10_cm']:.1f}cm / 90%ile {birth_thresholds['hc_p90_cm']:.1f}cm）"
            )
    else:
        hcL, hcM, hcS = birth_thresholds.get("hc_lms", (None, None, None))
        if hcL is not None and hcM is not None and hcS is not None:
            hc_z = value_to_lms_z(hcL, hcM, hcS, birth_head_circumference)
            hc_p = z_to_percentile(hc_z)
            hc_z_text = "-" if hc_z is None else f"{hc_z:+.2f}SD"
            hc_p_text = "-" if hc_p is None else f"{hc_p:.1f}%ile"
            st.markdown(
                f"頭囲: **{birth_head_circumference:.1f}cm / {hc_z_text} / {hc_p_text}**"
                f"（-2SD {birth_thresholds['hc_minus2sd_cm']:.1f}cm / 10%ile {birth_thresholds['hc_p10_cm']:.1f}cm / 90%ile {birth_thresholds['hc_p90_cm']:.1f}cm）"
            )
        else:
            st.markdown(
                f"頭囲: **{birth_head_circumference:.1f}cm** / データ参照値なし"
            )

    if birth_plane_fig is not None:
        st.plotly_chart(birth_plane_fig, width='stretch')

# 推奨事項の表示
st.subheader("✅ 管理のポイント")
specials = guidance.get('special_management', [])

for special in specials:
    is_needed = special.get('needed', True)
    if is_needed:
        st.markdown(f"**{special['title']}**")
        if special.get('title') == '💊 ケイツーシロップ12回投与法':
            if birth_date:
                d0 = birth_date.strftime('%Y/%m/%d')

                d1 = (birth_date + timedelta(days=1)).strftime('%Y/%m/%d')
                d4 = (birth_date + timedelta(days=4)).strftime('%Y/%m/%d')
                d11 = (birth_date + timedelta(days=11)).strftime('%Y/%m/%d')
            else:
                d0 = d1 = d4 = d11 = None

            third = special.get('k2_third_to_twelfth')
           
            b1, b2, b3 = st.columns(3)
            with b1:
                st.markdown("#### 1回目")
                st.caption(f"日齢0（{d0}） / 日齢1（{d1}）" if d0 else "日齢0 / 日齢1")
                st.markdown("- 点滴あり：日齢0に静注（ELBWは半量）")
                st.markdown("- 点滴なし：日齢1に内服（ELBWも減量しない）")

            with b2:
                st.markdown("#### 2回目")
                st.caption(f"日齢4（{d4}）" if d4 else "日齢4")
                st.markdown("- 消化不良：静注（ELBWは半量）")
                st.markdown("- 消化良好：内服")

            with b3:
                st.markdown("#### 3〜12回目")
                st.caption(f"日齢11（{d11}）以降" if d11 else "日齢11以降")
                if third:
                    st.markdown(f"- {third}")

                else:
                    st.markdown("- 日齢11以降の最初の水曜から毎週水曜に内服")

            st.markdown("- 入院中の内服は処置オーダで指示する")
            st.markdown("- 退院処方として12回目までを処方する")
            continue
        for item in special.get('items', []):
            st.markdown(item)
    else:
        st.markdown(
            f"<span style='color: gray;'><b>{special['title']}</b></span>",
            unsafe_allow_html=True
        )
        for item in special.get('items', []):
                       st.markdown(f"<span style='color: gray;'>{item}</span>", unsafe_allow_html=True)

st.markdown("---")
st.markdown("## 💡 光線療法基準")

st.markdown("### ✅ 現在の基準値")
sum1, sum2 = st.columns(2)
with sum1:
    st.markdown("**村田・井村の基準**")
    if phototherapy_category is None:
        st.markdown("適用ライン: **体重未測定のため判定不可**")
    else:
        st.markdown(f"適用ライン: **{phototherapy_category}**")
    st.markdown(f"日齢: {days_old}日")
    if phototherapy_threshold is None:
        st.metric("TB基準値", "未定義")
    else:
        st.metric("TB基準値", f"{phototherapy_threshold} mg/dL")

with sum2:
    st.markdown("**📊 神戸大学（森岡）の基準**")
    if morioka is None:
        st.markdown("対象外")
    else:
        pca_low, pca_high = morioka["pca_group"]
        if pca_high == float("inf"):
            pca_label = f"{pca_low}週以上"
        else:
            pca_label = f"{pca_low}-{pca_high}週" if pca_low != pca_high else f"{pca_low}週"

        st.markdown(f"修正週数: **{pca_label}**")
        st.markdown(f"出生後時間: **{morioka['time_label']}**（{hours_old:.1f}時間）")
        morioka_tb1, morioka_tb2, morioka_tb3 = st.columns(3)
        with morioka_tb1:
            st.metric("TB low", f"{morioka['tb']['low']} mg/dL")
        with morioka_tb2:
            st.metric("TB high", f"{morioka['tb']['high']} mg/dL")
        with morioka_tb3:
            st.metric("TB 交換輸血", f"{morioka['tb']['exchange']} mg/dL")
        st.markdown(f"UB（µg/dL） low/high/交換輸血: **{morioka['ub']['low']}/{morioka['ub']['high']}/{morioka['ub']['exchange']}**")

if days_old == 0:
    st.info("日齢0のため、神戸大学（森岡）の基準を参照してください。")

st.markdown("### 📈 村田・井村の基準")

if phototherapy_category is None:
    st.markdown("適用基準ライン: **体重未測定のため判定不可**")
else:
    st.markdown(f"適用基準ライン: **{phototherapy_category}**")

risk_factors = []
if apgar_score_5min <= 3:
    risk_factors.append("5分Apgar≦3")
if respiratory_distress:
    risk_factors.append("呼吸窮迫（PaO2≦40が2時間以上持続）")
if acidosis:
    risk_factors.append("アシドーシス（pH≦7.15）")
if hypothermia:
    risk_factors.append("低体温（直腸温<35℃が2時間以上持続）")
if hypoproteinemia:
    risk_factors.append("低蛋白血症（血清蛋白≦4.0またはAlb≦2.5）")
if hypoglycemia:
    risk_factors.append("低血糖")
if hemolysis:
    risk_factors.append("溶血")
if cns_abnormality:
    risk_factors.append("敗血症を含む中枢神経系の異常徴候")

if is_day0:
    st.caption("今日は0日目です。0日目は村田・井村の基準値が定義されていません。")

if adjusted:
    risk_factors_str = "、".join(risk_factors)
    st.caption(f"⚠️ 核黄疸危険因子（{risk_factors_str}）により基準を1段階低く調整（{original_category} → {phototherapy_category}）")
elif has_kernicterus_risk:
    risk_factors_str = "、".join(risk_factors)
    st.caption(f"⚠️ 核黄疸危険因子（{risk_factors_str}）あり（最低基準 {phototherapy_category} を適用）")
else:
    st.caption("✅ 核黄疸危険因子なし")

fig = go.Figure()

category_order = ["≥ 2,500g", "2,000 ~ 2,499g", "1,500 ~ 1,999g", "1,000 ~ 1,499g", "≤ 999g"]
colors = {
    "≥ 2,500g": "#1f77b4",
    "2,000 ~ 2,499g": "#ff7f0e",
    "1,500 ~ 1,999g": "#2ca02c",
    "1,000 ~ 1,499g": "#d62728",
    "≤ 999g": "#9467bd"
}

for cat in category_order:
    thresholds = MURATA_PHOTOTHERAPY_THRESHOLDS[cat]
    days = list(range(1, 8))
    values = [thresholds[d] for d in days]
    is_highlighted = (cat == phototherapy_category)

    fig.add_trace(go.Scatter(
        x=days,
        y=values,
        mode='lines+markers',
        name=cat,
        line=dict(
            width=4 if is_highlighted else 2,
            color=colors[cat],
            dash='solid' if is_highlighted else 'dot'
        ),
        marker=dict(
            size=8 if is_highlighted else 6,
            symbol='circle'
        )
    ))

if phototherapy_threshold is not None and not is_day0:
    x_today = 7 if days_old >= 7 else days_old
    day_label = "7以上" if days_old >= 7 else str(days_old)
    fig.add_hline(
        y=phototherapy_threshold,
        line_dash="dash",
        line_color=colors[phototherapy_category],
        line_width=2,
    )

    fig.add_trace(go.Scatter(
        x=[x_today],
        y=[phototherapy_threshold],
        mode="markers",
        name="今日の基準",
        marker=dict(
            size=16,
            color=colors[phototherapy_category],
            line=dict(color="white", width=2),
            symbol="circle",
        ),
        showlegend=False,
        hovertemplate=(
            f"日齢: {day_label}日<br>TB基準: {phototherapy_threshold} mg/dL<extra></extra>"
        ),
    ))

    fig.add_annotation(
        x=x_today,
        y=phototherapy_threshold,
        text=f"今日（日齢{day_label}日）\nTB基準 {phototherapy_threshold} mg/dL",
        showarrow=True,
        arrowhead=2,
        ax=40,
        ay=-40,
        font=dict(size=12, color="white"),
        bgcolor="rgba(0,0,0,0.55)",
        bordercolor="rgba(255,255,255,0.35)",
        borderwidth=1,
    )

fig.update_layout(
    xaxis_title="生後日齢（日）",
    yaxis_title="血清総ビリルビン値（mg/dL）",
    hovermode='x unified',
    height=500,
    legend=dict(
        yanchor="top",
        y=0.99,
        xanchor="left",
        x=0.01
    )
)

fig.update_xaxes(range=[0.5, 7.5], tickmode="linear", dtick=1)

st.plotly_chart(fig, width='stretch')

st.markdown("---")
st.markdown("### 📊 神戸大学（森岡）の基準")
if morioka is None:
    st.info("神戸大学（森岡）の基準は修正週数が22週以上の範囲で参照できます。")
else:
    pca_low, pca_high = morioka["pca_group"]
    if pca_high == float("inf"):
        pca_label = f"{pca_low}週以上"
    else:
        pca_label = f"{pca_low}-{pca_high}週" if pca_low != pca_high else f"{pca_low}週"

    headline = f"修正週数: **{pca_label}** / 出生後時間: **{morioka['time_label']}**（{hours_old:.1f}時間）"
    subline = "（表のTBは low/high/交換輸血 の順。UBは別途閾値。）"

    time_buckets = [24, 48, 72, 96, 120, float("inf")]
    bucket_ranges = [(0, 24), (24, 48), (48, 72), (72, 96), (96, 120), (120, float("inf"))]

    birth_total_days = gestational_weeks * 7 + gestational_days

    def corrected_weeks_at_day(day_int):
        return int((birth_total_days + int(day_int)) // 7)

    highlight_pairs = set()
    for b, (h_start, h_end) in zip(time_buckets, bucket_ranges):
        if h_end == float("inf"):
            min_day = int(h_start // 24)
            w0 = corrected_weeks_at_day(min_day)
            for g in MORIOKA_TB_THRESHOLDS.keys():
                gl, gh = g
                if gh == float("inf"):
                    if gl <= w0:
                        highlight_pairs.add((g, b))
                else:
                    if gh >= w0:
                        highlight_pairs.add((g, b))
        else:
            min_day = int(h_start // 24)
            max_day = int((h_end - 1e-9) // 24)
            for d in range(min_day, max_day + 1):
                g = get_morioka_pca_group_from_weeks(corrected_weeks_at_day(d))
                if g is not None:
                    highlight_pairs.add((g, b))

    if hours_old < 24:
        st.info("生後24時間未満のため、神戸大学（森岡）の基準は参考値です。")
    if corrected_weeks < 22:
        st.warning("修正週数が22週未満のため、神戸大学（森岡）の基準は参考値です。")

    morioka_table_html = build_morioka_html_table(
        current_pca_group=(pca_low, pca_high),
        current_time_bucket_hours=morioka["time_bucket_hours"],
        highlight_pairs=highlight_pairs,
    )

    st.markdown(headline)
    st.caption(subline)
    st.markdown(morioka_table_html, unsafe_allow_html=True)

    # UBの閾値もテキストで表示
    ub_low, ub_high, ub_ex = MORIOKA_UB_THRESHOLDS[(pca_low, pca_high)]
    st.markdown(f"**UB（µg/dL） low/high/交換輸血:** {ub_low}/{ub_high}/{ub_ex}")
